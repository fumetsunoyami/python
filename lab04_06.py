
from openpyxl import *
from tkinter import *

# globally declare wb and sheet variable

# opening the existing excel file
wb = load_workbook('C:\\Users\\vietl\\Downloads\\Book.xlsx')
# create the sheet object
sheet = wb.active


def excel():
	
	# resize the width of columns in
	# excel spreadsheet
	sheet.column_dimensions['A'].width = 10
	sheet.column_dimensions['B'].width = 40
	sheet.column_dimensions['C'].width = 12
	sheet.column_dimensions['D'].width = 40
	sheet.column_dimensions['E'].width = 12
	sheet.column_dimensions['F'].width = 6
	sheet.column_dimensions['G'].width = 6
	sheet.column_dimensions['H'].width = 50

	# write given data to an excel spreadsheet
	# at particular location
	sheet.cell(row=1, column=1).value = "Mssv"
	sheet.cell(row=1, column=2).value = "Họ tên"
	sheet.cell(row=1, column=3).value = "Ngày sinh"
	sheet.cell(row=1, column=4).value = "Email"
	sheet.cell(row=1, column=5).value = "Số đt"
	sheet.cell(row=1, column=6).value = "Học kỳ"
	sheet.cell(row=1, column=7).value = "Năm học"
	sheet.cell(row=1, column=8).value = "Môn học"


# Function to set focus (cursor)
def focus1(event):
	# set focus on the name_field box
	name_field.focus_set()
	


# Function to set focus
def focus2(event):
	# set focus on the ngaysinh_field box
	ngaysinh_field.focus_set()


# Function to set focus
def focus3(event):
	# set focus on the email_field box
	email_field.focus_set()


# Function to set focus
def focus4(event):
	# set focus on the contact_no_field box
	contact_no_field.focus_set()


# Function to set focus
def focus5(event):
	# set focus on the hocky_field box
	hocky_field.focus_set()


# Function to set focus
def focus6(event):
	# set focus on the namhoc_field box
	namhoc_field.focus_set()
def focus7(event):
	# set focus on the namhoc_field box
	monhoc_field.focus_set()


# Function for clearing the
# contents of text entry boxes
def clear():
	
	# clear the content of text entry box
	mssv_field.delete(0, END)
	name_field.delete(0, END)
	ngaysinh_field.delete(0, END)
	email_field.delete(0, END)
	contact_no_field.delete(0, END)
	hocky_field.delete(0, END)
	namhoc_field.delete(0, END)
	monhoc_field.delete(0, END)


# Function to take data from GUI
# window and write to an excel file
def insert():
	
	# if user not fill any entry
	# then print "empty input"
	if mssv_field.get() == "": 
		
			
		print("empty input")

	else:

		# assigning the max row and max column
		# value upto which data is written
		# in an excel sheet to the variable
		current_row = sheet.max_row
		current_column = sheet.max_column

		# get method returns current text
		# as string which we write into
		# excel spreadsheet at particular location
		sheet.cell(row=current_row + 1, column=1).value = name_field.get()
		sheet.cell(row=current_row + 1, column=2).value = name_field.get()
		sheet.cell(row=current_row + 1, column=3).value = ngaysinh_field.get()
		sheet.cell(row=current_row + 1, column=4).value = email_field.get()
		sheet.cell(row=current_row + 1, column=5).value = contact_no_field.get()
		sheet.cell(row=current_row + 1, column=6).value = hocky_field.get()
		sheet.cell(row=current_row + 1, column=7).value = namhoc_field.get()
		sheet.cell(row=current_row + 1, column=8).value = monhoc_field.get()

		# save the file
		wb.save('C:\\Users\\vietl\\Downloads\\Book.xlsx')

		# set focus on the name_field box
		name_field.focus_set()

		# call the clear() function
		clear()


# Driver code
if __name__ == "__main__":
	
	# create a GUI window
	root = Tk()

	# set the background colour of GUI window
	root.configure(background='light green')

	# set the title of GUI window
	root.title("registration form")

	# set the configuration of GUI window
	root.geometry("500x300")

	excel()

	# create a Form label
	heading = Label(root, text="Form", bg="light green")

	# create a Name label
	mssv = Label(root, text="mssv", bg="light green")

	# create a Course label
	name = Label(root, text="name", bg="light green")

	# create a Semester label
	ngaysinh = Label(root, text="ngay sinh", bg="light green")

	# create a Form No. label
	form_no = Label(root, text="email", bg="light green")

	# create a Contact No. label
	contact_no = Label(root, text="sdt", bg="light green")

	# create a Email id label
	email_id = Label(root, text="Học kỳ", bg="light green")

	# create a address label
	address = Label(root, text="Năm học", bg="light green")
	monhoc = Label(root, text="môn học", bg="light green")
	# grid method is used for placing
	# the widgets at respective positions
	# in table like structure .
	heading.grid(row=0, column=1)
	mssv.grid(row=1, column=0)
	name.grid(row=2, column=0)
	ngaysinh.grid(row=3, column=0)
	form_no.grid(row=4, column=0)
	contact_no.grid(row=5, column=0)
	email_id.grid(row=6, column=0)
	address.grid(row=7, column=0)
	monhoc.grid(row=8, column = 0)

	# create a text entry box
	# for typing the information
	mssv_field = Entry(root)
	name_field = Entry(root)
	ngaysinh_field = Entry(root)
	email_field = Entry(root)
	contact_no_field = Entry(root)
	hocky_field = Entry(root)
	namhoc_field = Entry(root)
	monhoc_field = Entry(root)

	# bind method of widget is used for
	# the binding the function with the events

	# whenever the enter key is pressed
	# then call the focus1 function
	mssv_field.bind("<Return>", focus1)

	# whenever the enter key is pressed
	# then call the focus2 function
	name_field.bind("<Return>", focus2)

	# whenever the enter key is pressed
	# then call the focus3 function
	ngaysinh_field.bind("<Return>", focus3)

	# whenever the enter key is pressed
	# then call the focus4 function
	email_field.bind("<Return>", focus4)

	# whenever the enter key is pressed
	# then call the focus5 function
	contact_no_field.bind("<Return>", focus5)

	# whenever the enter key is pressed
	# then call the focus6 function
	hocky_field.bind("<Return>", focus6)
 	# whenever the enter key is pressed
	# then call the focus7 function
	namhoc_field.bind("<Return>", focus7)

	# grid method is used for placing
	# the widgets at respective positions
	# in table like structure .
	name_field.grid(row=1, column=1, ipadx="100")
	name_field.grid(row=2, column=1, ipadx="100")
	ngaysinh_field.grid(row=3, column=1, ipadx="100")
	email_field.grid(row=4, column=1, ipadx="100")
	contact_no_field.grid(row=5, column=1, ipadx="100")
	hocky_field.grid(row=6, column=1, ipadx="100")
	namhoc_field.grid(row=7, column=1, ipadx="100")


	# call excel function
	excel()

	# create a Submit Button and place into the root window
	submit = Button(root, text="Submit", fg="Black",
							bg="Red", command=insert)
	submit.grid(row=8, column=1)

	# start the GUI
	root.mainloop()
